from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib.auth import authenticate, login, logout
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.db.models import Max

from smuckers.models import Bol, BolItem, ForkliftDriver, TruckDriver, SentEmail, Weight
from smuckers.forms import BolForm, BolItemForm, ForkliftDriverForm, TruckDriverForm, SentEmailForm, WeightForm

from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.core.mail import send_mail
from django.core.mail import EmailMessage

import math
import datetime

#admin:B00st123

#forklift_user
#guest123

#truck_user
#guest123

#manager_user
#guest123

# Where do forklift initials and signiture goes
# 30 lbs per pallet
# Add truck drivers


def index(request):
    return render(request, 'smuckers/index.html')

def display(request):
	bols = Bol.objects.filter(truck_approved=False).values()
	context = {'bols': bols}
	return render(request, 'smuckers/display.html', context)

def displayManager(request):
	bols = Bol.objects.filter(approved=False).filter(truck_approved=True).values()
	context = {'bols': bols}
	return render(request, 'smuckers/display.html', context)

def displayAll(request):
	bols = Bol.objects.all().values()
	context = {'bols': bols}
	return render(request, 'smuckers/display.html', context)

def displaybol(request, bol_id):
	bol = Bol.objects.get(id=bol_id)
	bolItems = BolItem.objects.filter(bol=bol).values()
	bolForm = BolForm(Bol.objects.filter(id=bol_id).values()[0])

	bolItemArray = []

	for bolItem in bolItems:
		bolItemArray.append({'item': bolItem,'form': BolItemForm(bolItem)})#, prefix=bolItem['id']))

	truckDrivers = TruckDriver.objects.all()

	context = {'bolItemsForm': bolItemArray, 'bol': bol, 'bolForm': bolForm, 'truckDrivers': truckDrivers}
	return render(request, 'smuckers/displaybol.html', context)

def approvebol(request, bol_id):
	bol = Bol.objects.get(id=bol_id)
	if request.user.groups.filter(name='Truck').exists():
		bol.truck_approved = True;
	else:
		bol.approved = True;
	bol.save();
	bols = Bol.objects.filter(approved=False).values()
	context = {'bols': bols}

	# Send Email
	saveFile(bol_id)
	sentEmails = SentEmail.objects.all()
	emails = []
	for email in sentEmails:
		emails.append(email.email)
	emailMessage = EmailMessage(
	    str(bol.bol_number) + ': ' + bol.date.strftime("%Y-%m-%d %H:%M"),
	    '',
	    'from@example.com',
	    emails
	)
	emailMessage.attach_file('smuckers/resources/downloads/download.xlsm')
	emailMessage.send()

	if request.user.groups.filter(name='Truck').exists():
		return redirect('/smuckers/display-data')
	return redirect('/smuckers/display-all')

def unapprovebol(request, bol_id):
	bol = Bol.objects.get(id=bol_id)
	if request.user.groups.filter(name='Truck').exists():
		bol.truck_approved = False;
	else:
		bol.approved = False;
		bol.truck_approved = False;
	bol.save();
	bols = Bol.objects.filter(approved=False).values()
	context = {'bols': bols}
	if request.user.groups.filter(name='Truck').exists():
		return redirect('/smuckers/display-data')
	return redirect('/smuckers/display-all')

def deletebol(request, bol_id):
	bol = Bol.objects.get(id=bol_id)
	bol.delete()
	bols = Bol.objects.filter(approved=False).values()
	context = {'bols': bols}

	if request.user.groups.filter(name='Truck').exists():
		return redirect('/smuckers/display-data')
	return redirect('/smuckers/display-all')

def downloadbol(request, bol_id):
	response = HttpResponse(saveFile(bol_id), content_type='application/vnd.ms-excel')
	return response

def enter(request):

	#if not request.user.groups.filter(name='Forklift').exists():


	if request.method == 'POST': # and request.FILES['inputfile']:
		# myfile = request.FILES['inputfile']
		# fs = FileSystemStorage()
		# filename = fs.save('smuckers/files/' + myfile.name, myfile)
		# uploaded_file_url = fs.url(filename)
		# f = open(uploaded_file_url, 'r')
		# f.readline()
		stringList = request.POST.get('inputfile', '').split('\n')
		# stringList = []
		# for line in f:
		# 	stringList.append(line)

		bol = Bol()

		bol_number_max = Bol.objects.all().aggregate(Max('bol_number'))
		if bol_number_max['bol_number__max'] is None:
			bol.bol_number = 110000
		else:
			bol.bol_number = bol_number_max['bol_number__max'] + 1

		bol.date = datetime.datetime.now()
		bol.dock_number = request.POST.get('dock_number', '')
		bol.seal_number = request.POST.get('seal_number', '')
		fdId = int(request.POST.get('forklift_driver', 0))
		bol.forklift_driver = ForkliftDriver.objects.get(id=fdId)
		bol.save()

		weights = Weight.objects.all()
		drum_weight_table = {}
		for weight in weights:
			weight_data = {}
			weight_data['no_drum'] = weight.without_drum
			weight_data['with_drum'] = weight.with_drum
			drum_weight_table[weight.item_number] = weight_data

		stringList = remove_invalid_scans(stringList)
		itemCode_date_unique_list = get_unique_drums(stringList)
		itemCode_date_dict = calculate_count_of_unique_drums(stringList)
		itemCode_date_dict = calculate_weight_of_unique_drums(itemCode_date_dict,drum_weight_table)

		for key in itemCode_date_dict:
			bolItem = BolItem()
			bolItem.bol = bol
			bolItem.description = 'Drums ' + key
			bolItem.packages = itemCode_date_dict[key]['count']
			bolItem.weight = itemCode_date_dict[key]['weight']
			bolItem.save()

		pallet_count = calculate_pallet_count(itemCode_date_dict)
		pallet_weight = calculate_pallet_weight(pallet_count)

		bolItem = BolItem()
		bolItem.bol = bol
		bolItem.description = 'Pallets'
		bolItem.packages = pallet_count
		bolItem.weight = pallet_weight
		bolItem.save()


	forkliftDrivers = ForkliftDriver.objects.all()
		#return render(request, 'smuckers/test.html', {'test': itemCode_date_dict}) 
		# return redirect('/smuckers/' + str(bol.id) + "/displaybol")
	return render(request, 'smuckers/enter.html', {'forkliftDrivers': forkliftDrivers})

def bolPost(request):
		bol_id = request.POST.get('id', '')
		bol = Bol.objects.get(id=bol_id)

		bol.bill_of_lading = request.POST.get('bill_of_lading', '')
		bol.shipping_order = request.POST.get('shipping_order', '')
		bol.shippers_number = request.POST.get('shippers_number', '')
		bol.agents_number = request.POST.get('agents_number', '')
		bol.date = request.POST.get('date', '')
		bol.from_address = request.POST.get('from_address', '')
		bol.cosigned_to = request.POST.get('cosigned_to', '')
		bol.destination = request.POST.get('destination', '')
		bol.state_of = request.POST.get('state_of', '')
		bol.county_of = request.POST.get('county_of', '')
		bol.route = request.POST.get('route', '')
		bol.delivering_carrier = request.POST.get('delivering_carrier', '')
		bol.car_initial = request.POST.get('car_initial', '')
		bol.car_number = request.POST.get('car_number', '')

		bol.save()

		return redirect('/smuckers/' + bol_id + "/displaybol")
		# return render(request, 'smuckers/test.html', {'test': request.POST.get('description', '')})

def bolItemPost(request):
	bol_item_id = request.POST.get('id', '')
	bol_id = request.POST.get('bol_id', '')
	bolItem = BolItem.objects.get(id=bol_item_id)

	bolItem.packages = request.POST.get('packages', '')
	bolItem.description = request.POST.get('description', '')
	bolItem.weight = request.POST.get('weight', '')
	bolItem.class_or_rate = request.POST.get('class_or_rate', '')
	if request.POST.get('check_column', '') == 'on':
		bolItem.check_column = True
	else:
		bolItem.check_column = False

	bolItem.save()

	return redirect('/smuckers/' + bol_id + "/displaybol")
	# return render(request, 'smuckers/test.html', {'test': test})

def excelTest(request):
	emailMessage = EmailMessage(
	    'Subject',
	    'Body',
	    'from@example.com',
	    ['dpu2010@gmail.com']
	)
	emailMessage.attach_file('smuckers/resources/downloads/download.xlsm')
	emailMessage.send()

	return render(request, 'smuckers/test.html')

def instructions(request):
	return render(request, 'smuckers/instructions.html')

def forkliftEntry(request):
	if request.method == 'POST':
		first_name = request.POST.get('first_name', '')
		last_name = request.POST.get('last_name', '')
		forkliftDriver = ForkliftDriver()
		forkliftDriver.first_name = first_name
		forkliftDriver.last_name = last_name
		forkliftDriver.save()

	context = {'forkliftForm': ForkliftDriverForm()}
	return render(request, 'smuckers/forkliftentry.html', context)

def truckEntry(request):
	if request.method == 'POST':
		first_name = request.POST.get('first_name', '')
		last_name = request.POST.get('last_name', '')
		truckDriver = TruckDriver()
		truckDriver.first_name = first_name
		truckDriver.last_name = last_name
		truckDriver.save()

	context = {'forkliftForm': ForkliftDriverForm()}
	return render(request, 'smuckers/truckentry.html', context)

def sentEmailEntry(request):
	if request.method == 'POST':
		email = request.POST.get('email', '')
		email_id = request.POST.get('sent_email', '')
		if email != '':
			sentEmail = SentEmail()
			sentEmail.email = email
			sentEmail.save()
		else:
			email_id = request.POST.get('sent_email', '')
			email = SentEmail.objects.get(id=email_id)
			email.delete()

	sentEmails = SentEmail.objects.all()
	context = {'sentEmailForm': SentEmailForm(), 'sentEmails': sentEmails}
	return render(request, 'smuckers/sentemail.html', context)

def weightEntry(request):
	if request.method == 'POST':
		email = request.POST.get('email', '')
		weight_id = request.POST.get('weight_id', '')
		if weight_id == '':
			weight = Weight()
			weight.item_number = request.POST.get('item_number', '')
			weight.with_drum = request.POST.get('with_drum', '')
			weight.without_drum = request.POST.get('without_drum', '')
			weight.save()
		else:
			weight = Weight.objects.get(id=weight_id)
			weight.delete()

	weights = Weight.objects.all()
	context = {'weightForm': WeightForm(), 'weights': weights}
	return render(request, 'smuckers/weight.html', context)
#################################################
#                                               #
#                  LOGIN VIEWS                  #
#                                               #
#################################################
def loginUser(request):
    return render(request, 'smuckers/login.html')

def loginPost(request):
	username = request.POST.get('username', '')
	password = request.POST.get('password', '')
	user = authenticate(username=username, password=password)
	if user is not None:
		login(request, user)
		if request.user.groups.filter(name='Forklift').exists():
			return redirect('/smuckers/enter-data')
		elif request.user.groups.filter(name='Truck').exists():
			return redirect('/smuckers/display-data')
		return redirect('/smuckers/display-all')
	else:
		return redirect('/smuckers/login')

def logoutUser(request):
	logout(request)
	return redirect('/smuckers/login')

def saveFile(bol_id):
	bol = Bol.objects.get(id=bol_id)
	bolItems = BolItem.objects.filter(bol=bol)

	wb = load_workbook(filename = 'smuckers/resources/boltemplate.xlsx')
	ws = wb["BOL"]
	ws['E5'] = bol.bill_of_lading
	ws['E7'] = bol.shipping_order
	ws['Q5'] = bol.shippers_number
	ws['P7'] = bol.agents_number
	ws['Q11'] = bol.date
	ws['C13'] = bol.from_address

	ws['D19'] = bol.cosigned_to
	ws['D21'] = bol.destination
	ws['I21'] = bol.state_of
	ws['N21'] = bol.county_of
	ws['C23'] = bol.route
	ws['E25'] = bol.delivering_carrier
	ws['J25'] = bol.car_initial
	ws['O25'] = bol.car_number
	ws['K53'] = bol.truck_driver.first_name + ' ' + bol.truck_driver.last_name

	ws['O38'] = bol.forklift_driver.first_name[0] + bol.forklift_driver.last_name[0]

	row = 28

	total_weight = 0

	for bolItem in bolItems:
		ws['A' + str(row)] = bolItem.packages
		ws['C' + str(row)] = bolItem.description
		ws['I' + str(row)] = bolItem.weight
		total_weight += bolItem.weight
		ws['K' + str(row)] = bolItem.class_or_rate
		if bolItem.check_column:
			ws['M' + str(row)] = 'YES'
		row += 1

	ws['D40'] = bol.seal_number
	ws['I40'] = str(total_weight)

	wb.save('smuckers/resources/downloads/download.xlsm')
	return save_virtual_workbook(wb)

# drum_weight_table = {
# 	'51663':{'no_drum': 384, 'with_drum': 409},
# 	'10834':{'no_drum': 390, 'with_drum': 415},
# 	'51359':{'no_drum': 384, 'with_drum': 409},
# 	'50798':{'no_drum': 384, 'with_drum': 409},
# 	'51581':{'no_drum': 340, 'with_drum': 365},
# 	'50688':{'no_drum': 340, 'with_drum': 365},
# 	'50629':{'no_drum': 340, 'with_drum': 365},
# 	'50628':{'no_drum': 340, 'with_drum': 365},
# 	'10449':{'no_drum': 384, 'with_drum': 409},
# 	'50670':{'no_drum': 384, 'with_drum': 409},
# 	'10854':{'no_drum': 452, 'with_drum': 477},
# 	'10523':{'no_drum': 340, 'with_drum': 365},
# 	'51559':{'no_drum': 340, 'with_drum': 365},
# 	'50552':{'no_drum': 384, 'with_drum': 409},
# 	'50290':{'no_drum': 384, 'with_drum': 409},
# 	'51325':{'no_drum': 340, 'with_drum': 365},
# 	'50551':{'no_drum': 384, 'with_drum': 409},
# 	'51874':{'no_drum': 384, 'with_drum': 409},
# 	'11170':{'no_drum': 384, 'with_drum': 409},
# 	'50668':{'no_drum': 340, 'with_drum': 365},
# 	'12015':{'no_drum': 384, 'with_drum': 409},
# 	'12012':{'no_drum': 384, 'with_drum': 409},
# 	'10504':{'no_drum': 384, 'with_drum': 409},
# 	'11185':{'no_drum': 452, 'with_drum': 477},
# 	'11367':{'no_drum': 452, 'with_drum': 477},
# 	'11150':{'no_drum': 452, 'with_drum': 477},
# 	'10254':{'no_drum': 340, 'with_drum': 365},
# 	'10840':{'no_drum': 452, 'with_drum': 477},
# 	'10122':{'no_drum': 452, 'with_drum': 477},
# 	'12020':{'no_drum': 370, 'with_drum': 395},
# 	'51358':{'no_drum': 384, 'with_drum': 409},
# 	'51676':{'no_drum': 384, 'with_drum': 409},
# 	'50687':{'no_drum': 340, 'with_drum': 365},
# 	'50686':{'no_drum': 340, 'with_drum': 365},
# 	'10757':{'no_drum': 340, 'with_drum': 365},
# 	'13129':{'no_drum': 384, 'with_drum': 409},
# 	'50674':{'no_drum': 384, 'with_drum': 409},
# 	'11352':{'no_drum': 384, 'with_drum': 409},
# 	'10530':{'no_drum': 452, 'with_drum': 477},
# 	'12075':{'no_drum': 452, 'with_drum': 477},
# 	'10258':{'no_drum': 452, 'with_drum': 477},
# 	'51692':{'no_drum': 452, 'with_drum': 477},
# 	'13249':{'no_drum': 452, 'with_drum': 477},
# 	'13157':{'no_drum': 452, 'with_drum': 477},
# 	'13114':{'no_drum': 384, 'with_drum': 409},
# 	'13265':{'no_drum': 452, 'with_drum': 477},
# 	'13139':{'no_drum': 452, 'with_drum': 477},
# 	'13481':{'no_drum': 452, 'with_drum': 477},
# 	'13153':{'no_drum': 360, 'with_drum': 385},
# 	'13277':{'no_drum': 452, 'with_drum': 477}
# }

#Straight Forward
def remove_invalid_scans(scan_list):
	# Replace letter, 9th char, with a "D" first
	scan_list = [scan[:8] + 'D' + scan[9:] for scan in scan_list]
	# Remove scans that are too short
	scan_list = [scan for scan in scan_list if len(scan) > 12]
	# Remove duplicate scans
	scan_list = list(set(scan_list))
	return scan_list

#center
def get_unique_drums(scan_list):
	# Drum = item code + date
	itemCode_date_list = [scan[3:15] for scan in scan_list]
	itemCode_date_unique_list = list(set(itemCode_date_list))
	return itemCode_date_unique_list

#left
def calculate_count_of_unique_drums(scan_list):
	itemCode_date_list = [scan[3:15] for scan in scan_list]
	# Calculate the counts for each drum
	itemCode_date_dict = {}
	for drum in itemCode_date_list:
		if drum in itemCode_date_dict:
			itemCode_date_dict[drum]['count'] += 1
		else:
			itemCode_date_dict[drum] = {'count': 1}
	return itemCode_date_dict

#right
def calculate_weight_of_unique_drums(itemCode_date_dict, drum_weight_table):
	for drum in itemCode_date_dict:
		itemCode = drum[:5]
		print('Item Code: ' + itemCode)
		count = itemCode_date_dict[drum]['count']
		with_drum = drum_weight_table[itemCode]['with_drum']
		itemCode_date_dict[drum]['weight'] = count * with_drum
	return itemCode_date_dict

#left
def calculate_pallet_count(itemCode_date_dict):
	total_drums = 0
	for drum in itemCode_date_dict:
		total_drums += itemCode_date_dict[drum]['count']
	pallet_count = math.ceil(total_drums/4)
	return pallet_count

#right
def calculate_pallet_weight(pallet_count):
	#30?
	pallet_weight = pallet_count * 30
	return pallet_weight

